import openpyxl
from openpyxl.chart import BarChart, Reference
import logging

# 1. Logging Setup (Txt file)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.FileHandler("openpyxl_output_log.txt", mode='w'), logging.StreamHandler()]
)
logger = logging.getLogger()

# 2. Create Source
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2025"
ws.append(["Product", "Quantity", "Price"])
ws.append(["A", 10, 50]); ws.append(["B", 5, 100]); ws.append(["C", 8, 75])
wb.save('openpyxl_source.xlsx')
logger.info("Source file 'openpyxl_source.xlsx' created.")

# 3. Process
wb = openpyxl.load_workbook('openpyxl_source.xlsx')
ws = wb['2025']
ws['D1'] = "Total"
for r in range(2, ws.max_row + 1):
    total = ws.cell(row=r, column=2).value * ws.cell(row=r, column=3).value
    ws.cell(row=r, column=4).value = total
    logger.info(f"Row {r}: Product {ws.cell(row=r, column=1).value} - Total calculated: {total}")

# 4. Embed Chart
chart = BarChart()
chart.title = "OpenPyXL Internal Chart"
data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "F2")

wb.save('openpyxl_summary.xlsx')
logger.info("Saved log to openpyxl_output_log.txt and chart inside openpyxl_summary.xlsx")